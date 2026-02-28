# backend/app/infrastructure/export/excel_exporter.py
from __future__ import annotations

import asyncio
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.entities.assumption import Assumption
from app.domain.entities.deal import Deal
from app.domain.interfaces.providers import ExcelExporter


class OpenpyxlExcelExporter(ExcelExporter):
    """Excel exporter using openpyxl to create .xlsx workbooks."""

    async def export(
        self, deal: Deal, assumptions: list[Assumption]
    ) -> bytes:
        return await asyncio.to_thread(
            self._build_workbook, deal, assumptions
        )

    @staticmethod
    def _build_workbook(
        deal: Deal, assumptions: list[Assumption]
    ) -> bytes:
        wb = Workbook()

        # Styles
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        # ---------------------------------------------------------------
        # Sheet 1: Deal Inputs
        # ---------------------------------------------------------------
        ws_deal = wb.active
        ws_deal.title = "Deal Inputs"

        deal_rows = [
            ("Field", "Value"),
            ("Deal Name", deal.name),
            ("Address", deal.address),
            ("City", deal.city),
            ("State", deal.state),
            ("Property Type", deal.property_type.value),
            ("Square Feet", deal.square_feet),
            ("Latitude", deal.latitude),
            ("Longitude", deal.longitude),
        ]

        for r, (field, value) in enumerate(deal_rows, start=1):
            ws_deal.cell(row=r, column=1, value=field)
            ws_deal.cell(row=r, column=2, value=value)

        # Format header row
        for col in range(1, 3):
            cell = ws_deal.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws_deal.column_dimensions["A"].width = 20
        ws_deal.column_dimensions["B"].width = 30

        # ---------------------------------------------------------------
        # Sheet 2: Assumptions
        # ---------------------------------------------------------------
        ws_assumptions = wb.create_sheet("Assumptions")

        assumption_headers = [
            "Key", "Value", "Unit", "Range Min", "Range Max", "Source", "Notes"
        ]
        for col, header in enumerate(assumption_headers, start=1):
            cell = ws_assumptions.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, a in enumerate(assumptions, start=2):
            ws_assumptions.cell(row=r, column=1, value=a.key)
            ws_assumptions.cell(row=r, column=2, value=a.value_number)
            ws_assumptions.cell(row=r, column=3, value=a.unit)
            ws_assumptions.cell(row=r, column=4, value=a.range_min)
            ws_assumptions.cell(row=r, column=5, value=a.range_max)
            ws_assumptions.cell(row=r, column=6, value=a.source_type.value)
            ws_assumptions.cell(row=r, column=7, value=a.notes)

        for col in range(1, 8):
            ws_assumptions.column_dimensions[get_column_letter(col)].width = 18

        # Write to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
